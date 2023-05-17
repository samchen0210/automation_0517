import os
import logging
import configparser
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from openpyxl import load_workbook

config = configparser.ConfigParser(interpolation=configparser.ExtendedInterpolation())
config.read('config.ini')

def open(driver):
    
    driver.get(config.get('web', 'url'))
    driver.maximize_window()
    time.sleep(2)

def write_res(data, save_mode, case_id):
    wb = load_workbook('Barco_result.xlsx')
    sheet = wb['Sheet1']

    if save_mode == 'response':
        sheet.cell(case_id + 4 , 6).value = str(data)
    elif save_mode == 'result':
        sheet.cell(case_id + 4 , 7).value = str(data)

    wb.save('Barco_result.xlsx')

